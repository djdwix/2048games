from flask import Flask, request, jsonify, send_from_directory, redirect, url_for
from flask_cors import CORS
import sqlite3
import random
import string
import os
import time
import threading
from functools import wraps
import socket
import hashlib
import json
import atexit
from datetime import datetime, timedelta
import psutil
import ssl
import subprocess
import re

app = Flask(__name__)
CORS(app)

PUBLIC_DIR = 'public'
DATA_DIR = 'data'
if not os.path.exists(PUBLIC_DIR):
    os.makedirs(PUBLIC_DIR)
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

DATABASE = 'phone_numbers.db'

CARRIER_PREFIXES = {
    '中国移动': ['134', '135', '136', '137', '138', '139', '147', '148', '150', 
                '151', '152', '157', '158', '159', '172', '178', '182', '183', 
                '184', '187', '188', '198', '195', '197'],
    '中国联通': ['130', '131', '132', '145', '146', '155', '156', '166', '171', 
                '175', '176', '185', '186', '196'],
    '中国电信': ['133', '149', '153', '173', '177', '180', '181', '189', '191', 
                '193', '199', '190', '197'],
    '中国广电': ['192', '194'],
    '虚拟运营商': ['170', '171', '165', '167', '162']
}

api_usage_stats = {
    'total_requests': 0,
    'successful_requests': 0,
    'failed_requests': 0,
    'endpoints': {},
    'hourly_stats': {}
}
api_stats_lock = threading.Lock()
system_start_time = time.time()

recent_requests = {}
request_lock = threading.Lock()

TC_CAPTCHA_APP_ID = '1314462072'

QUOTA_FILE = os.path.join(DATA_DIR, 'quota_data.json')
QUOTA_LIMIT = 30
QUOTA_LOCK = threading.Lock()

KEY_FILE = os.path.join(DATA_DIR, 'key_data.json')
KEY_LOCK = threading.Lock()

ENCRYPTION_KEY = 'd7a3f4c6e5b81290de4f3c2a1b0987654321fedcba0987654321abcdef567890'

ADMIN_TOKEN = '8TbmYBEnH3673'

def adapt_datetime(dt):
    return dt.isoformat()

def convert_datetime(text):
    return datetime.fromisoformat(text.decode())

sqlite3.register_adapter(datetime, adapt_datetime)
sqlite3.register_converter("TIMESTAMP", convert_datetime)

def encrypt_data(data):
    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    signature = hashlib.sha256((json_str + ENCRYPTION_KEY).encode()).hexdigest()
    return json.dumps({'data': data, 'signature': signature}, ensure_ascii=False)

def decrypt_and_verify(json_str):
    try:
        container = json.loads(json_str)
        data = container['data']
        signature = container['signature']
        
        data_json = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        expected_signature = hashlib.sha256((data_json + ENCRYPTION_KEY).encode()).hexdigest()
        
        if signature != expected_signature:
            return None
        return data
    except:
        return None

def load_quota_data():
    try:
        if not os.path.exists(QUOTA_FILE):
            return {}
        
        with open(QUOTA_FILE, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                return {}
            
            decrypted = decrypt_and_verify(content)
            if decrypted is None:
                backup_file = QUOTA_FILE + '.backup'
                if os.path.exists(backup_file):
                    with open(backup_file, 'r', encoding='utf-8') as backup_f:
                        backup_content = backup_f.read().strip()
                        if backup_content:
                            backup_decrypted = decrypt_and_verify(backup_content)
                            if backup_decrypted is not None:
                                with open(QUOTA_FILE, 'w', encoding='utf-8') as restore_f:
                                    restore_f.write(backup_content)
                                return backup_decrypted
                return {}
            
            with open(QUOTA_FILE + '.backup', 'w', encoding='utf-8') as backup_f:
                backup_f.write(content)
            
            return decrypted
    except Exception as e:
        return {}

def save_quota_data(data):
    try:
        encrypted_data = encrypt_data(data)
        with open(QUOTA_FILE, 'w', encoding='utf-8') as f:
            f.write(encrypted_data)
        with open(QUOTA_FILE + '.backup', 'w', encoding='utf-8') as backup_f:
            backup_f.write(encrypted_data)
        return True
    except Exception as e:
        return False

def load_key_data():
    try:
        if not os.path.exists(KEY_FILE):
            return {'keys': {}, 'requests': {}, 'current_hour': ''}
        
        with open(KEY_FILE, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                return {'keys': {}, 'requests': {}, 'current_hour': ''}
            
            decrypted = decrypt_and_verify(content)
            if decrypted is None:
                backup_file = KEY_FILE + '.backup'
                if os.path.exists(backup_file):
                    with open(backup_file, 'r', encoding='utf-8') as backup_f:
                        backup_content = backup_f.read().strip()
                        if backup_content:
                            backup_decrypted = decrypt_and_verify(backup_content)
                            if backup_decrypted is not None:
                                with open(KEY_FILE, 'w', encoding='utf-8') as restore_f:
                                    restore_f.write(backup_content)
                                return backup_decrypted
                return {'keys': {}, 'requests': {}, 'current_hour': ''}
            
            with open(KEY_FILE + '.backup', 'w', encoding='utf-8') as backup_f:
                backup_f.write(content)
            
            return decrypted
    except Exception as e:
        return {'keys': {}, 'requests': {}, 'current_hour': ''}

def save_key_data(data):
    try:
        encrypted_data = encrypt_data(data)
        with open(KEY_FILE, 'w', encoding='utf-8') as f:
            f.write(encrypted_data)
        with open(KEY_FILE + '.backup', 'w', encoding='utf-8') as backup_f:
            backup_f.write(encrypted_data)
        return True
    except Exception as e:
        return False

def cleanup_key_data():
    try:
        now = time.time()
        with KEY_LOCK:
            key_data = load_key_data()
            
            keys_to_delete = []
            for key, key_info in key_data.get('keys', {}).items():
                if now - key_info.get('created_at', 0) > 300:
                    keys_to_delete.append(key)
            
            for key in keys_to_delete:
                del key_data['keys'][key]
            
            save_key_data(key_data)
    except Exception as e:
        pass

def generate_key():
    characters = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(random.choices(characters, k=12))

def get_client_ip():
    if request.headers.get('X-Forwarded-For'):
        ip = request.headers['X-Forwarded-For'].split(',')[0]
    elif request.headers.get('X-Real-IP'):
        ip = request.headers['X-Real-IP']
    else:
        ip = request.remote_addr
    
    if ip == '127.0.0.1':
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(('8.8.8.8', 80))
            ip = s.getsockname()[0]
        except:
            pass
        finally:
            s.close()
    
    return ip

def check_quota(client_ip):
    now = datetime.now()
    hour_key = now.strftime('%Y-%m-%d %H:00')
    
    with QUOTA_LOCK:
        quota_data = load_quota_data()
        
        if hour_key not in quota_data:
            quota_data[hour_key] = {}
        
        if client_ip not in quota_data[hour_key]:
            quota_data[hour_key][client_ip] = 0
        
        if quota_data[hour_key][client_ip] >= QUOTA_LIMIT:
            return False, quota_data[hour_key][client_ip]
        
        quota_data[hour_key][client_ip] += 1
        save_quota_data(quota_data)
        
        return True, quota_data[hour_key][client_ip]

def cleanup_quota_data():
    try:
        now = datetime.now()
        with QUOTA_LOCK:
            quota_data = load_quota_data()
            
            keys_to_delete = []
            for hour_key in quota_data:
                hour_datetime = datetime.strptime(hour_key, '%Y-%m-%d %H:00')
                if now.date() != hour_datetime.date():
                    keys_to_delete.append(hour_key)
            
            for key in keys_to_delete:
                del quota_data[key]
            
            save_quota_data(quota_data)
    except Exception as e:
        pass

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(('10.255.255.255', 1))
            ip = s.getsockname()[0]
        except:
            ip = '127.0.0.1'
        finally:
            s.close()
        return ip

def get_server_ip():
    try:
        if request.headers.get('X-Forwarded-For'):
            ip = request.headers['X-Forwarded-For'].split(',')[0]
        elif request.headers.get('X-Real-IP'):
            ip = request.headers['X-Real-IP']
        else:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect(('8.8.8.8', 80))
                ip = s.getsockname()[0]
            except:
                ip = '127.0.0.1'
            finally:
                s.close()
        
        return ip
    except:
        return '127.0.0.1'

def get_carrier_by_prefix(phone_number):
    if len(phone_number) < 3:
        return '未知'
    
    prefix = phone_number[:3]
    for carrier, prefixes in CARRIER_PREFIXES.items():
        if prefix in prefixes:
            return carrier
    
    if prefix.startswith('19') and prefix in ['190', '191', '192', '193', '195', '196', '197']:
        return '卫星通信'
    
    return '未知'

def is_valid_custom_prefix(prefix):
    if len(prefix) != 3 or not prefix.isdigit():
        return False
    
    for carrier, prefixes in CARRIER_PREFIXES.items():
        if prefix in prefixes:
            return True
    
    if prefix.startswith('19') and prefix in ['190', '191', '192', '193', '195', '196', '197']:
        return True
    
    return False

def generate_security_code():
    characters = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(random.choices(characters, k=6))

def generate_phone_number(custom_prefix=None):
    all_prefixes = []
    for carrier, prefixes in CARRIER_PREFIXES.items():
        all_prefixes.extend(prefixes)
    
    all_prefixes.extend(['190', '191', '192', '193', '195', '196', '197'])
    
    if custom_prefix and len(custom_prefix) == 3 and custom_prefix.isdigit():
        if not is_valid_custom_prefix(custom_prefix):
            raise ValueError("自定义号段无效或未被分配运营商")
        prefix = custom_prefix
    else:
        prefix = random.choice(all_prefixes)
    
    middle = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    suffix = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    return prefix + middle + suffix

def get_db_connection():
    conn = sqlite3.connect(DATABASE, detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute('PRAGMA cache_size=20000')
    conn.execute('PRAGMA temp_store=MEMORY')
    conn.execute('PRAGMA mmap_size=268435456')
    
    return conn

def parse_datetime(dt_str):
    if not dt_str:
        return None
    
    try:
        if isinstance(dt_str, datetime):
            return dt_str
            
        if '.' in dt_str:
            dt_str = dt_str.split('.')[0]
        
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d %H:%M:%S.%f'
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue
        
        return None
    except Exception:
        return None

def is_maintenance_time():
    now = datetime.now()
    hour = now.hour
    minute = now.minute
    second = now.second
    
    if hour == 23 and minute >= 50:
        return True
    if hour == 0 and minute < 10:
        return True
    if hour == 0 and minute == 9 and second > 0:
        return True
    if hour == 0 and minute == 10 and second == 0:
        return False
    
    return False

def should_show_maintenance_page():
    return is_maintenance_time()

def init_database():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS phone_numbers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        phone_number TEXT UNIQUE NOT NULL,
        security_code TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        is_used INTEGER DEFAULT 0,
        last_accessed TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        client_ip TEXT,
        carrier TEXT,
        category TEXT DEFAULT '通用',
        purpose TEXT,
        access_count INTEGER DEFAULT 0,
        request_id TEXT,
        custom_prefix TEXT,
        security_code_generated_at TIMESTAMP,
        security_code_expires_at TIMESTAMP
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS usage_stats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date DATE NOT NULL,
        total_requests INTEGER DEFAULT 0,
        successful_generations INTEGER DEFAULT 0,
        failed_requests INTEGER DEFAULT 0,
        unique_clients INTEGER DEFAULT 0,
        UNIQUE(date)
    )
    ''')
    
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_phone_number ON phone_numbers(phone_number)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_security_code ON phone_numbers(security_code)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_is_used ON phone_numbers(is_used)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_created_at ON phone_numbers(created_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_client_ip ON phone_numbers(client_ip)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_category ON phone_numbers(category)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_request_id ON phone_numbers(request_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_custom_prefix ON phone_numbers(custom_prefix)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_security_code_expires_at ON phone_numbers(security_code_expires_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_comp_used_created ON phone_numbers(is_used, created_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_comp_ip_used ON phone_numbers(client_ip, is_used)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_comp_code_expires ON phone_numbers(security_code_expires_at, is_used)')
    
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_usage_stats_date ON usage_stats(date)')
    
    cursor.execute('PRAGMA optimize')
    
    conn.commit()
    conn.close()

def cleanup_unused_phone_numbers():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            DELETE FROM phone_numbers 
            WHERE (is_used = 0 AND security_code IS NULL AND 
                  (strftime('%s', 'now') - strftime('%s', created_at)) > 300)
               OR (security_code IS NOT NULL AND is_used = 0 AND 
                  (strftime('%s', 'now') - strftime('%s', last_accessed)) > 3600)
        ''')
        
        conn.commit()
        conn.close()
    except Exception as e:
        pass

def update_usage_stats(success=True):
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM usage_stats WHERE date = ?', (today,))
        existing = cursor.fetchone()
        
        if not existing:
            cursor.execute('''
                INSERT INTO usage_stats (date, total_requests, successful_generations, 
                                        failed_requests, unique_clients)
                VALUES (?, ?, ?, ?, ?)
            ''', (today, 1, 1 if success else 0, 0 if success else 1, 0))
        else:
            cursor.execute('UPDATE usage_stats SET total_requests = total_requests + 1 WHERE date = ?', (today,))
            
            if success:
                cursor.execute('UPDATE usage_stats SET successful_generations = successful_generations + 1 WHERE date = ?', (today,))
            else:
                cursor.execute('UPDATE usage_stats SET failed_requests = failed_requests + 1 WHERE date = ?', (today,))
        
        conn.commit()
        conn.close()
    except Exception as e:
        pass

def update_api_stats(endpoint, success=True):
    with api_stats_lock:
        api_usage_stats['total_requests'] += 1
        if success:
            api_usage_stats['successful_requests'] += 1
        else:
            api_usage_stats['failed_requests'] += 1
        
        hour = datetime.now().strftime('%Y-%m-%d %H:00')
        if hour not in api_usage_stats['hourly_stats']:
            api_usage_stats['hourly_stats'][hour] = {'total': 0, 'success': 0, 'failed': 0}
        
        api_usage_stats['hourly_stats'][hour]['total'] += 1
        if success:
            api_usage_stats['hourly_stats'][hour]['success'] += 1
        else:
            api_usage_stats['hourly_stats'][hour]['failed'] += 1
        
        if endpoint not in api_usage_stats['endpoints']:
            api_usage_stats['endpoints'][endpoint] = {'total': 0, 'success': 0, 'failed': 0}
        
        api_usage_stats['endpoints'][endpoint]['total'] += 1
        if success:
            api_usage_stats['endpoints'][endpoint]['success'] += 1
        else:
            api_usage_stats['endpoints'][endpoint]['failed'] += 1

def rate_limit():
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            client_ip = get_client_ip()
            now = time.time()
            
            with request_lock:
                if client_ip in recent_requests:
                    last_request_time = recent_requests[client_ip]
                    time_diff = now - last_request_time
                    min_wait = 5.00
                    if time_diff < min_wait:
                        wait_time = max(min_wait - time_diff, 0)
                        update_api_stats(request.endpoint, False)
                        return jsonify({
                            'success': False,
                            'error': f'请求过于频繁，请等待{wait_time:.1f}秒后再试',
                            'cooldown': wait_time
                        }), 429
                
                recent_requests[client_ip] = now
            
            cleanup_threshold = now - 30
            with request_lock:
                to_delete = []
                for ip, timestamp in recent_requests.items():
                    if timestamp < cleanup_threshold:
                        to_delete.append(ip)
                for ip in to_delete:
                    del recent_requests[ip]
            
            try:
                response = f(*args, **kwargs)
                success = True
                if isinstance(response, tuple):
                    success = response[1] < 400
                update_api_stats(request.endpoint, success)
                return response
            except Exception as e:
                update_api_stats(request.endpoint, False)
                raise e
        return decorated_function
    return decorator

@app.before_request
def before_request():
    now = time.time()
    if now - getattr(app, 'last_cleanup', 0) > 60:
        cleanup_unused_phone_numbers()
        cleanup_quota_data()
        cleanup_key_data()
        app.last_cleanup = now
    
    if request.path.startswith('/api/'):
        return
    
    if is_maintenance_time():
        if request.path != '/404.html':
            return send_from_directory(PUBLIC_DIR, '404.html'), 503
    else:
        if request.path == '/404.html':
            if not request.referrer or '404.html' not in request.referrer:
                return redirect(url_for('index'))

@app.route('/')
def index():
    if is_maintenance_time():
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    return send_from_directory(PUBLIC_DIR, 'index.html')

@app.route('/quota-key', methods=['GET'])
def quota_key_page():
    if is_maintenance_time():
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    return send_from_directory(PUBLIC_DIR, 'quota-key.html')

@app.route('/quota', methods=['GET'])
def quota_management():
    try:
        if is_maintenance_time():
            return send_from_directory(PUBLIC_DIR, '404.html'), 503
        
        key = request.args.get('key', '').strip()
        client_ip = get_client_ip()
        
        if not key:
            return redirect(url_for('quota_key_page'))
        
        with KEY_LOCK:
            key_data = load_key_data()
            
            if key not in key_data.get('keys', {}):
                return redirect(url_for('quota_key_page'))
            
            key_info = key_data['keys'][key]
            
            if key_info['ip'] != client_ip:
                return redirect(url_for('quota_key_page'))
            
            if time.time() - key_info['created_at'] > 300:
                return redirect(url_for('quota_key_page'))
            
            return send_from_directory(PUBLIC_DIR, 'quota.html')
    except Exception as e:
        return redirect(url_for('quota_key_page'))

@app.route('/privacy-policy')
def privacy_policy():
    if is_maintenance_time():
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    return send_from_directory(PUBLIC_DIR, 'privacy-policy.html')

@app.route('/terms')
def terms_of_service():
    if is_maintenance_time():
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    return send_from_directory(PUBLIC_DIR, 'terms.html')

@app.route('/404.html')
def maintenance_page():
    return send_from_directory(PUBLIC_DIR, '404.html')

@app.route('/<path:path>')
def serve_static(path):
    if path == 'index.html' and is_maintenance_time():
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    
    if is_maintenance_time() and path not in ['404.html', 'style.css']:
        return send_from_directory(PUBLIC_DIR, '404.html'), 503
    
    return send_from_directory(PUBLIC_DIR, path)

def check_key_quota(client_ip):
    now = datetime.now()
    hour_key = now.strftime('%Y-%m-%d %H:00')
    
    with KEY_LOCK:
        key_data = load_key_data()
        
        if key_data['current_hour'] != hour_key:
            key_data['requests'] = {}
            key_data['current_hour'] = hour_key
        
        client_requests = key_data.get('requests', {}).get(client_ip, [])
        
        return len(client_requests) < 3

@app.route('/api/key/request', methods=['POST'])
def request_key():
    try:
        if is_maintenance_time():
            update_api_stats('request_key', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        client_ip = get_client_ip()
        
        if not check_key_quota(client_ip):
            update_api_stats('request_key', False)
            return jsonify({
                'success': False,
                'error': f'当前时段已申请3次密钥，请等待下一小时段'
            }), 429
        
        time.sleep(random.randint(15, 30))
        
        with KEY_LOCK:
            key_data = load_key_data()
            
            now = datetime.now()
            hour_key = now.strftime('%Y-%m-%d %H:00')
            
            if key_data['current_hour'] != hour_key:
                key_data['requests'] = {}
                key_data['current_hour'] = hour_key
                key_data['keys'] = {}
            
            new_key = generate_key()
            created_at = time.time()
            
            key_data['keys'][new_key] = {
                'ip': client_ip,
                'created_at': created_at,
                'used': False
            }
            
            if 'requests' not in key_data:
                key_data['requests'] = {}
            
            if client_ip not in key_data['requests']:
                key_data['requests'][client_ip] = []
            
            key_data['requests'][client_ip].append({
                'timestamp': created_at,
                'key': new_key
            })
            
            save_key_data(key_data)
            
            update_api_stats('request_key', True)
            
            return jsonify({
                'success': True,
                'key': new_key,
                'expires_in': 300
            })
    except Exception as e:
        update_api_stats('request_key', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/key/verify', methods=['POST'])
def verify_key():
    try:
        if is_maintenance_time():
            update_api_stats('verify_key', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        data = request.json or {}
        key = data.get('key', '').strip()
        client_ip = get_client_ip()
        
        if not key:
            update_api_stats('verify_key', False)
            return jsonify({
                'success': False,
                'error': '密钥不能为空'
            }), 400
        
        with KEY_LOCK:
            key_data = load_key_data()
            
            if key not in key_data.get('keys', {}):
                update_api_stats('verify_key', False)
                return jsonify({
                    'success': False,
                    'error': '密钥不存在或已过期'
                }), 404
            
            key_info = key_data['keys'][key]
            
            if key_info['ip'] != client_ip:
                update_api_stats('verify_key', False)
                return jsonify({
                    'success': False,
                    'error': '密钥与客户端IP不匹配'
                }), 403
            
            if time.time() - key_info['created_at'] > 300:
                update_api_stats('verify_key', False)
                return jsonify({
                    'success': False,
                    'error': '密钥已过期'
                }), 400
            
            update_api_stats('verify_key', True)
            return jsonify({
                'success': True,
                'message': '密钥验证成功',
                'key': key
            })
    except Exception as e:
        update_api_stats('verify_key', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/key/request-count', methods=['GET'])
def get_key_request_count():
    try:
        if is_maintenance_time():
            update_api_stats('get_key_request_count', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        client_ip = get_client_ip()
        
        with KEY_LOCK:
            key_data = load_key_data()
            
            now = datetime.now()
            hour_key = now.strftime('%Y-%m-%d %H:00')
            
            if key_data['current_hour'] != hour_key:
                request_count = 0
            else:
                client_requests = key_data.get('requests', {}).get(client_ip, [])
                request_count = len(client_requests)
            
            max_requests = 3
            
            next_hour = datetime(now.year, now.month, now.day, now.hour + 1, 0, 0)
            reset_time = time.mktime(next_hour.timetuple())
            
            update_api_stats('get_key_request_count', True)
            
            return jsonify({
                'success': True,
                'client_ip': client_ip,
                'request_count': request_count,
                'max_requests': max_requests,
                'remaining_requests': max_requests - request_count,
                'reset_time': reset_time
            })
    except Exception as e:
        update_api_stats('get_key_request_count', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as count FROM phone_numbers')
        total_records = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM phone_numbers WHERE is_used = 1')
        used_records = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(DISTINCT client_ip) as unique_clients FROM phone_numbers WHERE created_at >= datetime("now", "-1 day")')
        daily_clients = cursor.fetchone()['unique_clients']
        
        cursor.execute('SELECT COUNT(*) as recent_requests FROM phone_numbers WHERE created_at >= datetime("now", "-1 hour")')
        hourly_requests = cursor.fetchone()['recent_requests']
        
        conn.close()
        
        uptime = time.time() - system_start_time
        uptime_hours = uptime / 3600
        
        memory_info = psutil.virtual_memory()
        disk_usage = psutil.disk_usage('.')
        cpu_percent = psutil.cpu_percent(interval=0.1)
        
        now = datetime.now()
        recent_hour = now.strftime('%Y-%m-%d %H:00')
        hourly_data = api_usage_stats['hourly_stats'].get(recent_hour, {'total': 0, 'success': 0, 'failed': 0})
        
        maintenance_mode = is_maintenance_time()
        
        health_status = {
            'status': 'healthy' if not maintenance_mode else 'maintenance',
            'maintenance_mode': maintenance_mode,
            'database': {
                'total_records': total_records,
                'used_records': used_records,
                'connection': 'ok',
                'size_mb': os.path.getsize(DATABASE) / (1024 * 1024) if os.path.exists(DATABASE) else 0
            },
            'system': {
                'uptime_hours': round(uptime_hours, 2),
                'memory_usage_percent': memory_info.percent,
                'disk_usage_percent': disk_usage.percent,
                'cpu_usage_percent': cpu_percent,
                'active_threads': threading.active_count()
            },
            'network': {
                'local_ip': get_local_ip(),
                'server_ip': get_server_ip()
            },
            'performance': {
                'recent_hour_requests': hourly_data['total'],
                'recent_hour_success_rate': round((hourly_data['success'] / hourly_data['total'] * 100) if hourly_data['total'] > 0 else 100, 1),
                'daily_unique_clients': daily_clients,
                'hourly_requests': hourly_requests
            },
            'timestamps': {
                'server_time': datetime.now().isoformat(),
                'start_time': datetime.fromtimestamp(system_start_time).isoformat()
            }
        }
        
        return jsonify({
            'success': True,
            'health': health_status
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ip-info', methods=['GET'])
def get_ip_info():
    try:
        if is_maintenance_time():
            update_api_stats('get_ip_info', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        update_api_stats('get_ip_info', True)
        server_ip = get_server_ip()
        client_ip = get_client_ip()
        
        return jsonify({
            'success': True,
            'server_ip': server_ip,
            'client_ip': client_ip
        })
    except Exception as e:
        update_api_stats('get_ip_info', False)
        return jsonify({
            'success': False,
            'error': str(e),
            'server_ip': '获取失败',
            'client_ip': '获取失败'
        }), 500

@app.route('/api/custom/generate', methods=['POST'])
@rate_limit()
def generate_custom_number():
    try:
        if is_maintenance_time():
            update_usage_stats(False)
            update_api_stats('generate_custom_number', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        client_ip = get_client_ip()
        
        quota_result, current_count = check_quota(client_ip)
        if not quota_result:
            update_usage_stats(False)
            update_api_stats('generate_custom_number', False)
            return jsonify({
                'success': False,
                'error': f'配额已用尽（当前时段已使用{current_count}/{QUOTA_LIMIT}次），请稍后再试'
            }), 429
        
        data = request.json or {}
        custom_prefix = data.get('custom_prefix', '').strip()
        purpose = data.get('purpose', '')
        
        if custom_prefix and (len(custom_prefix) != 3 or not custom_prefix.isdigit()):
            update_usage_stats(False)
            update_api_stats('generate_custom_number', False)
            return jsonify({
                'success': False,
                'error': '自定义号段必须是3位数字'
            }), 400
        
        if custom_prefix and not is_valid_custom_prefix(custom_prefix):
            update_usage_stats(False)
            update_api_stats('generate_custom_number', False)
            return jsonify({
                'success': False,
                'error': '自定义号段无效或未被分配运营商，请使用有效的号段'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        max_attempts = 10
        for _ in range(max_attempts):
            try:
                phone_number = generate_phone_number(custom_prefix if custom_prefix else None)
            except ValueError as e:
                update_usage_stats(False)
                update_api_stats('generate_custom_number', False)
                return jsonify({
                    'success': False,
                    'error': str(e)
                }), 400
            
            carrier = get_carrier_by_prefix(phone_number)
            request_id = hashlib.md5(f"{phone_number}{client_ip}{time.time()}".encode()).hexdigest()[:16]
            
            cursor.execute(
                'SELECT id FROM phone_numbers WHERE phone_number = ?',
                (phone_number,)
            )
            existing = cursor.fetchone()
            
            if existing:
                continue
            
            try:
                cursor.execute(
                    '''INSERT INTO phone_numbers (phone_number, client_ip, carrier, category, purpose, request_id, custom_prefix) 
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (phone_number, client_ip, carrier, '通用', purpose, request_id, custom_prefix if custom_prefix else None)
                )
                conn.commit()
                
                update_usage_stats(True)
                update_api_stats('generate_custom_number', True)
                
                return jsonify({
                    'success': True,
                    'phone_number': phone_number,
                    'masked_phone': f"{phone_number[:3]}****{phone_number[7:]}",
                    'carrier': carrier,
                    'custom_prefix': custom_prefix if custom_prefix else '随机',
                    'purpose': purpose,
                    'request_id': request_id,
                    'quota_used': current_count,
                    'quota_remaining': QUOTA_LIMIT - current_count
                })
            except sqlite3.IntegrityError:
                continue
        
        update_usage_stats(False)
        update_api_stats('generate_custom_number', False)
        return jsonify({
            'success': False,
            'error': '生成失败，请重试'
        }), 500
    except Exception as e:
        update_usage_stats(False)
        update_api_stats('generate_custom_number', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/generate-code', methods=['POST'])
def generate_security_code_for_number():
    try:
        if is_maintenance_time():
            update_api_stats('generate_security_code_for_number', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        data = request.json
        phone_number = data.get('phone_number', '').strip()
        client_ip = get_client_ip()
        
        if not phone_number or len(phone_number) != 11:
            update_api_stats('generate_security_code_for_number', False)
            return jsonify({
                'success': False,
                'error': '手机号格式不正确'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'SELECT id, security_code, is_used, client_ip, access_count, security_code_generated_at, security_code_expires_at FROM phone_numbers WHERE phone_number = ?',
            (phone_number,)
        )
        
        result = cursor.fetchone()
        
        if not result:
            update_api_stats('generate_security_code_for_number', False)
            return jsonify({
                'success': False,
                'error': '手机号不存在'
            }), 404
        
        if result['is_used']:
            update_api_stats('generate_security_code_for_number', False)
            return jsonify({
                'success': False,
                'error': '该手机号已被使用'
            }), 400
        
        now = datetime.now()
        security_code_generated_at = None
        security_code_expires_at = None
        
        if result['security_code_generated_at']:
            security_code_generated_at = parse_datetime(result['security_code_generated_at'])
        
        if result['security_code_expires_at']:
            security_code_expires_at = parse_datetime(result['security_code_expires_at'])
        
        if result['security_code']:
            if security_code_expires_at and now < security_code_expires_at:
                if result['client_ip'] != client_ip:
                    update_api_stats('generate_security_code_for_number', False)
                    return jsonify({
                        'success': False,
                        'error': '安全码只能在生成时的客户端下使用'
                    }), 403
                update_api_stats('generate_security_code_for_number', False)
                return jsonify({
                    'success': False,
                    'error': '该手机号已有有效的安全码'
                }), 400
        
        security_code = generate_security_code()
        new_access_count = result['access_count'] + 1
        generated_at = now
        expires_at = now + timedelta(seconds=180)
        
        cursor.execute(
            '''UPDATE phone_numbers SET security_code = ?, last_accessed = CURRENT_TIMESTAMP, 
               access_count = ?, security_code_generated_at = ?, security_code_expires_at = ? WHERE phone_number = ?''',
            (security_code, new_access_count, generated_at, expires_at, phone_number)
        )
        conn.commit()
        
        update_api_stats('generate_security_code_for_number', True)
    
        return jsonify({
            'success': True,
            'security_code': security_code,
            'expires_in': 180,
            'generated_at': generated_at.isoformat(),
            'expires_at': expires_at.isoformat()
        })
    except Exception as e:
        update_api_stats('generate_security_code_for_number', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/check-code-expiry', methods=['POST'])
def check_security_code_expiry():
    try:
        if is_maintenance_time():
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        data = request.json
        phone_number = data.get('phone_number', '').strip()
        
        if not phone_number or len(phone_number) != 11:
            return jsonify({
                'success': False,
                'error': '手机号格式不正确'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'SELECT security_code, security_code_expires_at, is_used FROM phone_numbers WHERE phone_number = ?',
            (phone_number,)
        )
        
        result = cursor.fetchone()
        
        if not result:
            return jsonify({
                'success': False,
                'error': '手机号不存在'
            }), 404
        
        if result['is_used']:
            return jsonify({
                'success': False,
                'error': '该手机号已被使用',
                'is_expired': True,
                'can_regenerate': False
            }), 200
        
        if not result['security_code']:
            return jsonify({
                'success': True,
                'is_valid': False,
                'is_expired': True,
                'can_regenerate': True
            }), 200
        
        now = datetime.now()
        expires_at = None
        
        if result['security_code_expires_at']:
            expires_at = parse_datetime(result['security_code_expires_at'])
        
        is_valid = False
        is_expired = True
        remaining_seconds = 0
        
        if expires_at:
            if now < expires_at:
                is_valid = True
                is_expired = False
                remaining_seconds = int((expires_at - now).total_seconds())
            else:
                is_valid = False
                is_expired = True
                remaining_seconds = 0
        
        return jsonify({
            'success': True,
            'is_valid': is_valid,
            'is_expired': is_expired,
            'remaining_seconds': remaining_seconds,
            'can_regenerate': not is_valid
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/verify-copy', methods=['POST'])
def verify_and_copy():
    try:
        if is_maintenance_time():
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        data = request.json
        phone_number = data.get('phone_number', '').strip()
        security_code = data.get('security_code', '').strip().upper()
        captcha_ticket = data.get('captcha_ticket', '')
        captcha_randstr = data.get('captcha_randstr', '')
        client_ip = get_client_ip()
        
        if not captcha_ticket or not captcha_randstr:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '请先完成人机验证'
            }), 400
        
        if not phone_number or len(phone_number) != 11:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '手机号格式不正确'
            }), 400
        
        if not security_code or len(security_code) != 6:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '安全码格式不正确'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            'SELECT security_code, is_used, client_ip, security_code_expires_at FROM phone_numbers WHERE phone_number = ?',
            (phone_number,)
        )
        
        result = cursor.fetchone()
        
        if not result:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '手机号不存在'
            }), 404
        
        if result['is_used']:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '该手机号已被使用'
            }), 400
        
        if not result['security_code']:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '请先生成安全码'
            }), 400
        
        now = datetime.now()
        expires_at = None
        
        if result['security_code_expires_at']:
            expires_at = parse_datetime(result['security_code_expires_at'])
        
        if expires_at and now > expires_at:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '安全码已过期，请重新生成'
            }), 400
        
        if result['client_ip'] != client_ip:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '安全码只能在生成时的客户端下使用'
            }), 403
        
        if result['security_code'] != security_code:
            update_api_stats('verify_and_copy', False)
            return jsonify({
                'success': False,
                'error': '安全码错误'
            }), 400
        
        cursor.execute(
            'UPDATE phone_numbers SET is_used = 1, last_accessed = CURRENT_TIMESTAMP WHERE phone_number = ?',
            (phone_number,)
        )
        conn.commit()
        
        update_api_stats('verify_and_copy', True)
    
        return jsonify({
            'success': True,
            'phone_number': phone_number,
            'message': '验证成功，可以复制完整号码'
        })
    except Exception as e:
        update_api_stats('verify_and_copy', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/quota', methods=['GET'])
def get_quota_info():
    try:
        if is_maintenance_time():
            update_api_stats('get_quota_info', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        client_ip = get_client_ip()
        update_api_stats('get_quota_info', True)
        
        now = datetime.now()
        today_date = now.strftime('%Y-%m-%d')
        
        with QUOTA_LOCK:
            quota_data = load_quota_data()
            
            last_24h_stats = []
            for hour in range(24):
                hour_key = f"{today_date} {hour:02d}:00"
                hour_data = quota_data.get(hour_key, {})
                hour_total = sum(hour_data.values())
                hour_user_count = hour_data.get(client_ip, 0)
                
                last_24h_stats.append({
                    'hour': hour_key,
                    'hour_display': f"{hour:02d}:00",
                    'total_requests': hour_total,
                    'user_requests': hour_user_count,
                    'other_users': hour_total - hour_user_count
                })
            
            current_hour = now.strftime('%Y-%m-%d %H:00')
            current_count = quota_data.get(current_hour, {}).get(client_ip, 0)
            
            top_users = []
            current_hour_data = quota_data.get(current_hour, {})
            sorted_users = sorted(current_hour_data.items(), key=lambda x: x[1], reverse=True)[:10]
            
            for ip, count in sorted_users:
                top_users.append({
                    'ip': ip,
                    'count': count,
                    'is_current': ip == client_ip
                })
        
        return jsonify({
            'success': True,
            'quota': {
                'current_hour': current_hour,
                'client_ip': client_ip,
                'used': current_count,
                'limit': QUOTA_LIMIT,
                'remaining': QUOTA_LIMIT - current_count,
                'last_24h': last_24h_stats,
                'top_users': top_users
            }
        })
    except Exception as e:
        update_api_stats('get_quota_info', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/quota', methods=['GET'])
def get_all_quota():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': '未授权访问'}), 401
        
        token = auth_header.split(' ')[1]
        expected_token_hash = hashlib.sha256(ADMIN_TOKEN.encode()).hexdigest()
        
        if token != expected_token_hash:
            return jsonify({'success': False, 'error': '无效的授权令牌'}), 403
        
        with QUOTA_LOCK:
            quota_data = load_quota_data()
            
            summary = {}
            total_requests = 0
            unique_ips = set()
            
            for hour_key, hour_data in quota_data.items():
                hour_total = sum(hour_data.values())
                hour_ips = len(hour_data)
                summary[hour_key] = {
                    'total_requests': hour_total,
                    'unique_ips': hour_ips,
                    'average_per_ip': hour_total / hour_ips if hour_ips > 0 else 0
                }
                total_requests += hour_total
                unique_ips.update(hour_data.keys())
            
            overall_stats = {
                'total_hours': len(summary),
                'total_requests': total_requests,
                'total_unique_ips': len(unique_ips),
                'average_per_hour': total_requests / len(summary) if summary else 0,
                'average_per_ip': total_requests / len(unique_ips) if unique_ips else 0
            }
        
        update_api_stats('get_all_quota', True)
        return jsonify({
            'success': True,
            'quota_data': quota_data,
            'summary': summary,
            'overall_stats': overall_stats
        })
    except Exception as e:
        update_api_stats('get_all_quota', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/download-data', methods=['GET'])
def download_quota_data():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': '未授权访问'}), 401
        
        token = auth_header.split(' ')[1]
        expected_token_hash = hashlib.sha256(ADMIN_TOKEN.encode()).hexdigest()
        
        if token != expected_token_hash:
            return jsonify({'success': False, 'error': '无效的授权令牌'}), 403
        
        with QUOTA_LOCK:
            quota_data = load_quota_data()
        
        now = datetime.now()
        today_date = now.strftime('%Y-%m-%d')
        
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "配额数据"
            
            ws['A1'] = '时间段'
            ws['B1'] = '客户端IP'
            ws['C1'] = '请求次数'
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ['A1', 'B1', 'C1']:
                ws[cell].font = header_font
                ws[cell].fill = header_fill
                ws[cell].alignment = header_alignment
            
            row = 2
            for hour_key, hour_data in sorted(quota_data.items()):
                if hour_key.startswith(today_date):
                    for ip, count in hour_data.items():
                        ws[f'A{row}'] = hour_key
                        ws[f'B{row}'] = ip
                        ws[f'C{row}'] = count
                        row += 1
            
            for column in ['A', 'B', 'C']:
                max_length = 0
                column_letter = column
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            from flask import make_response
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=quota_data_{today_date}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            update_api_stats('download_quota_data', True)
            return response
        except ImportError:
            update_api_stats('download_quota_data', False)
            return jsonify({
                'success': False,
                'error': 'openpyxl 库未安装，无法生成Excel文件'
            }), 500
    except Exception as e:
        update_api_stats('download_quota_data', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/stats', methods=['GET'])
def get_stats():
    try:
        if is_maintenance_time():
            update_api_stats('get_stats', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        update_api_stats('get_stats', True)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as total FROM phone_numbers')
        total = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as used FROM phone_numbers WHERE is_used = 1')
        used = cursor.fetchone()['used']
        
        cursor.execute('SELECT COUNT(*) as available FROM phone_numbers WHERE is_used = 0')
        available = cursor.fetchone()['available']
        
        cursor.execute('SELECT COUNT(DISTINCT client_ip) as unique_clients FROM phone_numbers')
        unique_clients = cursor.fetchone()['unique_clients']
        
        cursor.execute('SELECT category, COUNT(*) as count FROM phone_numbers GROUP BY category')
        category_stats = {}
        for row in cursor.fetchall():
            category_stats[row['category']] = row['count']
        
        cursor.execute('SELECT SUM(total_requests) as total_reqs, SUM(successful_generations) as success_reqs, SUM(failed_requests) as failed_reqs FROM usage_stats')
        daily_stats = cursor.fetchone()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total,
                'used': used,
                'available': available,
                'unique_clients': unique_clients,
                'category_stats': category_stats,
                'daily_requests': daily_stats['total_reqs'] or 0,
                'daily_success': daily_stats['success_reqs'] or 0,
                'daily_failed': daily_stats['failed_reqs'] or 0,
                'api_stats': {
                    'total_requests': api_usage_stats['total_requests'],
                    'successful_requests': api_usage_stats['successful_requests'],
                    'failed_requests': api_usage_stats['failed_requests'],
                    'endpoint_stats': api_usage_stats['endpoints']
                }
            }
        })
    except Exception as e:
        update_api_stats('get_stats', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()

@app.route('/api/client-info', methods=['GET'])
def get_client_info():
    try:
        if is_maintenance_time():
            update_api_stats('get_client_info', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        update_api_stats('get_client_info', True)
        client_ip = get_client_ip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as total FROM phone_numbers WHERE client_ip = ?', (client_ip,))
        total_generated = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as used FROM phone_numbers WHERE client_ip = ? AND is_used = 1', (client_ip,))
        total_used = cursor.fetchone()['used']
        
        conn.close()
        
        return jsonify({
            'success': True,
            'client_ip': client_ip,
            'stats': {
                'total_generated': total_generated,
                'total_used': total_used,
                'available': total_generated - total_used
            }
        })
    except Exception as e:
        update_api_stats('get_client_info', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/api-stats', methods=['GET'])
def get_api_stats():
    try:
        if is_maintenance_time():
            update_api_stats('get_api_stats', False)
            return jsonify({
                'success': False,
                'error': '系统维护中，请稍后再试'
            }), 503
        
        update_api_stats('get_api_stats', True)
        hourly_stats_list = []
        for hour, stats in sorted(api_usage_stats['hourly_stats'].items(), reverse=True)[:24]:
            hourly_stats_list.append({
                'hour': hour,
                'total': stats['total'],
                'success': stats['success'],
                'failed': stats['failed'],
                'success_rate': round((stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0, 1)
            })
        
        endpoint_stats_list = []
        for endpoint, stats in api_usage_stats['endpoints'].items():
            endpoint_stats_list.append({
                'endpoint': endpoint,
                'total': stats['total'],
                'success': stats['success'],
                'failed': stats['failed'],
                'success_rate': round((stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0, 1)
            })
        
        return jsonify({
            'success': True,
            'api_stats': {
                'total_requests': api_usage_stats['total_requests'],
                'successful_requests': api_usage_stats['successful_requests'],
                'failed_requests': api_usage_stats['failed_requests'],
                'success_rate': round((api_usage_stats['successful_requests'] / api_usage_stats['total_requests'] * 100) if api_usage_stats['total_requests'] > 0 else 0, 2),
                'hourly_stats': hourly_stats_list,
                'endpoint_stats': endpoint_stats_list,
                'system_uptime_hours': round((time.time() - system_start_time) / 3600, 2)
            }
        })
    except Exception as e:
        update_api_stats('get_api_stats', False)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def cleanup_on_shutdown():
    cleanup_unused_phone_numbers()
    cleanup_quota_data()
    cleanup_key_data()

if __name__ == '__main__':
    try:
        import psutil
        psutil_available = True
    except ImportError:
        psutil_available = False
        psutil = None
    
    if not os.path.exists(PUBLIC_DIR):
        os.makedirs(PUBLIC_DIR)
    
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
    
    required_files = ['index.html', 'privacy-policy.html', 'terms.html', 'quota.html', 'quota-key.html', 'style.css', 'script.js', '404.html']
    for file_name in required_files:
        file_path = os.path.join(PUBLIC_DIR, file_name)
        if not os.path.exists(file_path):
            if os.path.exists(file_name):
                import shutil
                shutil.copy(file_name, file_path)
    
    init_database()
    cleanup_unused_phone_numbers()
    cleanup_quota_data()
    cleanup_key_data()
    
    atexit.register(cleanup_on_shutdown)
    
    import signal
    
    def signal_handler(signum, frame):
        cleanup_on_shutdown()
        exit(0)
    
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    host_ip = get_local_ip()
    port = 5000
    
    CERT_DIR = 'certs'
    KEY_FILE = 'private.key'
    CERT_FILE = 'certificate.crt'
    key_path = os.path.join(CERT_DIR, KEY_FILE)
    cert_path = os.path.join(CERT_DIR, CERT_FILE)
    
    if os.path.exists(key_path) and os.path.exists(cert_path):
        ssl_context = (cert_path, key_path)
        print("=" * 60)
        print("虚拟手机号生成器服务器启动成功!")
        print("=" * 60)
        print(f"HTTPS本地访问: https://localhost:{port}")
        print(f"HTTPS网络访问: https://unrepentant12.cloud:{port}")
        print(f"隐私政策: https://unrepentant12.cloud:{port}/privacy-policy")
        print(f"用户协议: https://unrepentant12.cloud:{port}/terms")
        print(f"配额密钥: https://unrepentant12.cloud:{port}/quota-key")
        print(f"健康检查: https://unrepentant12.cloud:{port}/api/health")
        print(f"API统计: https://unrepentant12.cloud:{port}/api/api-stats")
        print(f"数据目录: {DATA_DIR}")
        print(f"证书目录: {CERT_DIR}")
        print(f"密钥文件: {KEY_FILE}")
        print(f"证书文件: {CERT_FILE}")
        print(f"配额限制: 每小时{QUOTA_LIMIT}次生成")
        print(f"管理员令牌哈希: {hashlib.sha256(ADMIN_TOKEN.encode()).hexdigest()}")
        print(f"安全码有效期: 180秒")
        print(f"维护时间: 每日 23:50 - 00:10")
        print("=" * 60)
        
        app.last_cleanup = time.time()
        app.run(debug=True, host='0.0.0.0', port=port, threaded=True, ssl_context=ssl_context)
    else:
        print("=" * 60)
        print("证书文件未找到!")
        print("=" * 60)
        print("将启动HTTP服务器...")
        print(f"维护时间: 每日 23:50 - 00:10")
        print("=" * 60)
        
        app.last_cleanup = time.time()
        app.run(debug=True, host='0.0.0.0', port=port, threaded=True)